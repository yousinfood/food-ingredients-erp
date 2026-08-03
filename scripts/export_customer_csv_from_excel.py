#!/usr/bin/env python3
"""Export「客戶資料」from data/有信ERP.xlsx to CSV for local sync smoke test."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

CUSTOMER_SHEET = "客戶資料"
CUSTOMER_HEADERS = (
    "客戶編號",
    "區域",
    "客戶名稱",
    "聯絡人",
    "📞",
    "📞",
    "📞",
    "配送地址",
    "發票地址",
    "📍",
    "🟩Line",
    "付款方式",
    "固定配送日",
    "配送順序",
    "信用額度",
    "最後交易日",
    "備註",
)

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "有信ERP.xlsx"
OUT = ROOT / "data" / "exports" / "customers_local_test.csv"


def main() -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[CUSTOMER_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        if not rows:
            writer.writerow(list(CUSTOMER_HEADERS))
        else:
            for row in rows:
                writer.writerow(list(row) if row else [])
    print(OUT)


if __name__ == "__main__":
    main()
