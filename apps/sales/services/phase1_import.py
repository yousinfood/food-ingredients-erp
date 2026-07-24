from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook

from apps.inventory.models import Batch, Product, StockMovement
from apps.inventory.packaging import parse_packaging_spec
from apps.procurement.models import GoodsReceipt, PurchaseOrder, PurchaseOrderItem
from apps.production.models import ProductionOrder, Recipe, RecipeItem
from apps.sales.models import Customer, SalesOrder, SalesOrderItem

DEFAULT_EXCEL_PATH = settings.BASE_DIR / "data" / "有信ERP.xlsx"
CUSTOMER_SHEET = "客戶資料"
FINISHED_SHEET = "產品資料"
RAW_SHEET = "原料資料"

CUS_S001_DUPLICATE_NAME = "蔥油餅"
CUS_S001_REMAP_CODE = "CUS-S096"

CUSTOMER_HEADERS = (
    "客戶編號",
    "區域",
    "客戶名稱",
    "聯絡人",
    "📞",
    "📞",
    "📞",
    "配送地址",
    "發票地址",
    "📍",
    "🟩Line",
    "付款方式",
    "固定配送日",
    "配送順序",
    "信用額度",
    "最後交易日",
    "備註",
)

UNIT_MAP = {
    "包": Product.Unit.BOX,
    "箱": Product.Unit.BOX,
    "kg": Product.Unit.KG,
    "公斤": Product.Unit.KG,
    "g": Product.Unit.G,
    "公克": Product.Unit.G,
    "l": Product.Unit.L,
    "公升": Product.Unit.L,
    "ml": Product.Unit.ML,
    "毫升": Product.Unit.ML,
    "pcs": Product.Unit.PCS,
    "件": Product.Unit.PCS,
}


@dataclass
class ImportReport:
    customers: list[dict] = field(default_factory=list)
    finished_products: list[dict] = field(default_factory=list)
    raw_materials: list[dict] = field(default_factory=list)
    skipped_rows: list[dict] = field(default_factory=list)
    transformations: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    cleared: dict = field(default_factory=dict)
    imported: dict = field(default_factory=dict)

    @property
    def passed(self) -> bool:
        return not self.errors


def _norm(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (datetime, date)):
        return value
    s = str(value).strip()
    return s if s else None


def _phone_str(value) -> str:
    v = _norm(value)
    if v is None:
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    s = str(v).strip()
    while s and s[0] in {"'", "’", "/", "＇"}:
        s = s[1:].strip()
    return s


def _parse_date(value) -> date | None:
    v = _norm(value)
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(str(v), fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value) -> Decimal | None:
    v = _norm(value)
    if v is None:
        return None
    try:
        return Decimal(str(v).replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _parse_int(value) -> int | None:
    v = _norm(value)
    if v is None:
        return None
    try:
        return int(Decimal(str(v)))
    except (InvalidOperation, ValueError):
        return None


def _checkmark(value) -> bool:
    return _norm(value) in {"✓", "v", "V", "Y", "yes", "是", "1", "True", "true"}


def _packaging_defaults(*, spec: str, unit: str) -> dict:
    parsed = parse_packaging_spec(spec)
    if parsed:
        return {
            "sales_unit": parsed.get("sales_unit", Product.SalesUnit.PACK),
            "net_weight_value": parsed.get("net_weight_value"),
            "net_weight_unit": parsed.get("net_weight_unit", ""),
        }
    if unit == Product.Unit.KG:
        return {"sales_unit": Product.SalesUnit.KG}
    if unit == Product.Unit.BOX:
        return {"sales_unit": Product.SalesUnit.PACK}
    return {"sales_unit": Product.SalesUnit.PACK}


def _map_unit(value, default=Product.Unit.KG) -> str:
    raw = _norm(value)
    if raw is None:
        return default
    return UNIT_MAP.get(str(raw).lower(), UNIT_MAP.get(str(raw), default))


def _validate_customer_headers(headers: list) -> None:
    expected = list(CUSTOMER_HEADERS)
    actual = [_norm(h) or "" for h in headers[: len(expected)]]
    for i, label in enumerate(expected):
        if i >= len(actual):
            raise ValueError(f"「{CUSTOMER_SHEET}」缺少欄位：{label}")
        if actual[i] != label and not (label == "📞" and actual[i] == "📞"):
            raise ValueError(
                f"「{CUSTOMER_SHEET}」欄位不符：第 {i + 1} 欄預期「{label}」，實際「{actual[i]}」"
            )


def _resolve_customer_code(code: str, name: str, row_number: int, seen_codes: set[str], report: ImportReport) -> str | None:
    if not code:
        return None
    if code == "CUS-S001" and name == CUS_S001_DUPLICATE_NAME:
        if "CUS-S001" in seen_codes:
            report.transformations.append(
                {
                    "sheet": CUSTOMER_SHEET,
                    "row": row_number,
                    "field": "客戶編號",
                    "from": "CUS-S001",
                    "to": CUS_S001_REMAP_CODE,
                    "reason": f"重複代碼；保留 紅葡萄 為 CUS-S001，{name} 改用 {CUS_S001_REMAP_CODE}",
                }
            )
            return CUS_S001_REMAP_CODE
    if code in seen_codes:
        report.errors.append(f"第 {row_number} 列：客戶編號 {code} 重複（未處理）")
        return None
    seen_codes.add(code if code != "CUS-S001" or name != CUS_S001_DUPLICATE_NAME else code)
    if code == CUS_S001_REMAP_CODE:
        seen_codes.add(CUS_S001_REMAP_CODE)
    return code


def parse_customers(sheet, report: ImportReport) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        report.errors.append(f"「{CUSTOMER_SHEET}」工作表為空")
        return []

    headers = rows[0]
    try:
        _validate_customer_headers(list(headers))
    except ValueError as exc:
        report.errors.append(str(exc))
        return []

    seen_codes: set[str] = set()
    customers = []
    cus_s001_seen = False

    for row_number, row in enumerate(rows[1:], start=2):
        cells = list(row) + [None] * max(0, 17 - len(row))
        code = _norm(cells[0])
        name = _norm(cells[2])
        if not code and not name:
            continue
        if not code:
            report.skipped_rows.append({"sheet": CUSTOMER_SHEET, "row": row_number, "reason": "缺少客戶編號"})
            continue
        if not name:
            report.skipped_rows.append({"sheet": CUSTOMER_SHEET, "row": row_number, "reason": "缺少客戶名稱"})
            continue

        resolved_code = code
        if code == "CUS-S001" and name == CUS_S001_DUPLICATE_NAME and cus_s001_seen:
            resolved_code = CUS_S001_REMAP_CODE
            report.transformations.append(
                {
                    "sheet": CUSTOMER_SHEET,
                    "row": row_number,
                    "field": "客戶編號",
                    "from": "CUS-S001",
                    "to": CUS_S001_REMAP_CODE,
                    "reason": f"重複代碼；保留 紅葡萄 為 CUS-S001，{name} 改用 {CUS_S001_REMAP_CODE}",
                }
            )
        elif code == "CUS-S001":
            cus_s001_seen = True

        if resolved_code in seen_codes:
            report.errors.append(f"第 {row_number} 列：客戶編號 {resolved_code} 重複")
            continue
        seen_codes.add(resolved_code)

        phone = _phone_str(cells[4])
        phone_2 = _phone_str(cells[5])
        phone_3 = _phone_str(cells[6])
        credit = _parse_decimal(cells[14])
        last_tx = _parse_date(cells[15])
        delivery_seq = _parse_int(cells[13])

        record = {
            "row_number": row_number,
            "code": resolved_code,
            "name": name,
            "region": _norm(cells[1]) or "",
            "contact_person": _norm(cells[3]) or "",
            "phone": phone,
            "phone_2": phone_2,
            "phone_3": phone_3,
            "address": _norm(cells[7]) or "",
            "invoice_address": _norm(cells[8]) or "",
            "map_location": _norm(cells[9]) or "",
            "line_id": _norm(cells[10]) or "",
            "payment_method": _norm(cells[11]) or "",
            "delivery_day": _norm(cells[12]) or "",
            "delivery_sequence": delivery_seq,
            "credit_limit": credit,
            "last_transaction_date": last_tx.isoformat() if last_tx else None,
            "notes": _norm(cells[16]) or "",
            "is_active": True,
        }
        customers.append(record)

    return customers


def parse_finished_products(sheet, report: ImportReport) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        report.errors.append(f"「{FINISHED_SHEET}」工作表為空")
        return []

    products = []
    seen_skus: set[str] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        cells = list(row) + [None] * 11
        sku = _norm(cells[0])
        name = _norm(cells[1])
        if not sku and not name:
            continue
        if sku and not name:
            report.skipped_rows.append(
                {"sheet": FINISHED_SHEET, "row": row_number, "reason": f"料號 {sku} 缺少品名（placeholder）"}
            )
            continue
        if not sku:
            report.skipped_rows.append({"sheet": FINISHED_SHEET, "row": row_number, "reason": "缺少產品編號"})
            continue

        if sku in seen_skus:
            report.errors.append(f"「{FINISHED_SHEET}」第 {row_number} 列：產品編號 {sku} 重複")
            continue
        seen_skus.add(sku)

        can_be_raw = _checkmark(cells[6])
        product_kind = Product.ProductKind.DUAL if can_be_raw else Product.ProductKind.FINISHED
        unit = _map_unit(cells[4], default=Product.Unit.BOX)

        products.append(
            {
                "row_number": row_number,
                "sku": sku,
                "name": name,
                "product_kind": product_kind,
                "category": _norm(cells[2]) or "",
                "spec": _norm(cells[3]) or "",
                "unit": unit,
                "is_for_sale": _checkmark(cells[5]),
                "can_be_raw_material": can_be_raw,
                "is_active": _checkmark(cells[7]) if _norm(cells[7]) is not None else True,
                "notes": _norm(cells[8]) or "",
            }
        )
    return products


def parse_raw_materials(sheet, report: ImportReport) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        report.errors.append(f"「{RAW_SHEET}」工作表為空")
        return []

    materials = []
    seen_skus: set[str] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        cells = list(row) + [None] * 10
        sku = _norm(cells[0])
        name = _norm(cells[1])
        if not sku and not name:
            continue
        if not sku:
            report.skipped_rows.append({"sheet": RAW_SHEET, "row": row_number, "reason": "缺少原料編號"})
            continue
        if not name:
            report.skipped_rows.append({"sheet": RAW_SHEET, "row": row_number, "reason": "缺少原料名稱"})
            continue
        if sku in seen_skus:
            report.errors.append(f"「{RAW_SHEET}」第 {row_number} 列：原料編號 {sku} 重複")
            continue
        seen_skus.add(sku)

        unit_cost = _parse_decimal(cells[4])
        is_active = _checkmark(cells[8]) if _norm(cells[8]) is not None else True

        materials.append(
            {
                "row_number": row_number,
                "sku": sku,
                "name": name,
                "product_kind": Product.ProductKind.RAW,
                "category": _norm(cells[2]) or "",
                "unit": _map_unit(cells[3], default=Product.Unit.KG),
                "unit_cost": str(unit_cost) if unit_cost is not None else None,
                "is_for_sale": False,
                "can_be_raw_material": True,
                "is_active": is_active,
                "notes": _norm(cells[9]) or "",
            }
        )
    return materials


def run_preflight(excel_path: Path | None = None) -> ImportReport:
    path = Path(excel_path) if excel_path else DEFAULT_EXCEL_PATH
    report = ImportReport()

    if not path.exists():
        report.errors.append(f"找不到 Excel 檔案：{path}")
        return report

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in (CUSTOMER_SHEET, FINISHED_SHEET, RAW_SHEET):
        if sheet_name not in wb.sheetnames:
            report.errors.append(f"找不到工作表：{sheet_name}")

    if report.errors:
        wb.close()
        return report

    report.customers = parse_customers(wb[CUSTOMER_SHEET], report)
    report.finished_products = parse_finished_products(wb[FINISHED_SHEET], report)
    report.raw_materials = parse_raw_materials(wb[RAW_SHEET], report)
    wb.close()

    if len(report.customers) != 95:
        report.warnings.append(f"客戶列數為 {len(report.customers)}，預期 95（含 CUS-S096 轉換）")
    if len(report.finished_products) != 14:
        report.warnings.append(f"成品列數為 {len(report.finished_products)}，預期 14")
    if len(report.raw_materials) != 20:
        report.warnings.append(
            f"原料可匯入列數為 {len(report.raw_materials)}（Excel 共 25 列料號，"
            "其中 RM0014 缺名稱、RM0022–RM0025 為空白 placeholder）"
        )

    sku_overlap = {p["sku"] for p in report.finished_products} & {m["sku"] for m in report.raw_materials}
    if sku_overlap:
        report.errors.append(f"成品與原料料號衝突：{sorted(sku_overlap)}")

    return report


def clear_demo_master_data() -> dict:
    counts = {
        "sales_order_items": SalesOrderItem.objects.count(),
        "sales_orders": SalesOrder.objects.count(),
        "stock_movements": StockMovement.objects.count(),
        "batches": Batch.objects.count(),
        "production_orders": ProductionOrder.objects.count(),
        "recipe_items": RecipeItem.objects.count(),
        "recipes": Recipe.objects.count(),
        "purchase_order_items": PurchaseOrderItem.objects.count(),
        "goods_receipts": GoodsReceipt.objects.count(),
        "purchase_orders": PurchaseOrder.objects.count(),
        "products": Product.objects.count(),
        "customers": Customer.objects.count(),
    }

    SalesOrderItem.objects.all().delete()
    SalesOrder.objects.all().delete()
    StockMovement.objects.all().delete()
    Batch.objects.all().delete()
    ProductionOrder.objects.all().delete()
    RecipeItem.objects.all().delete()
    Recipe.objects.all().delete()
    PurchaseOrderItem.objects.all().delete()
    GoodsReceipt.objects.all().delete()
    PurchaseOrder.objects.all().delete()
    Product.objects.all().delete()
    Customer.objects.all().delete()

    return counts


@transaction.atomic
def execute_import(report: ImportReport) -> ImportReport:
    if not report.passed:
        report.errors.append("Preflight 未通過，取消匯入")
        return report

    report.cleared = clear_demo_master_data()

    for row in report.customers:
        Customer.objects.create(
            code=row["code"],
            name=row["name"],
            region=row["region"],
            contact_person=row["contact_person"],
            phone=row["phone"],
            phone_2=row["phone_2"],
            phone_3=row["phone_3"],
            address=row["address"],
            invoice_address=row["invoice_address"],
            map_location=row["map_location"],
            line_id=row["line_id"],
            payment_method=row["payment_method"],
            delivery_day=row["delivery_day"],
            delivery_sequence=row["delivery_sequence"],
            credit_limit=Decimal(row["credit_limit"]) if row["credit_limit"] else None,
            last_transaction_date=(
                date.fromisoformat(row["last_transaction_date"]) if row["last_transaction_date"] else None
            ),
            notes=row["notes"],
            is_active=row["is_active"],
        )

    for row in report.finished_products:
        Product.objects.create(
            sku=row["sku"],
            name=row["name"],
            product_kind=row["product_kind"],
            category=row["category"],
            spec=row["spec"],
            unit=row["unit"],
            is_for_sale=row["is_for_sale"],
            can_be_raw_material=row["can_be_raw_material"],
            is_active=row["is_active"],
            description=row["notes"],
            **_packaging_defaults(spec=row["spec"], unit=row["unit"]),
        )

    for row in report.raw_materials:
        Product.objects.create(
            sku=row["sku"],
            name=row["name"],
            product_kind=row["product_kind"],
            category=row["category"],
            unit=row["unit"],
            unit_cost=Decimal(row["unit_cost"]) if row["unit_cost"] else None,
            is_for_sale=False,
            can_be_raw_material=True,
            is_active=row["is_active"],
            description=row["notes"],
            **_packaging_defaults(spec="", unit=row["unit"]),
        )

    report.imported = {
        "customers": len(report.customers),
        "finished_products": len(report.finished_products),
        "raw_materials": len(report.raw_materials),
    }
    return report
