from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook

from apps.sales.models import Customer

DEFAULT_EXCEL_PATH = settings.BASE_DIR / "data" / "有信ERP.xlsx"
CUSTOMER_SHEET_NAME = "客戶資料"

COLUMN_ALIASES = {
    "code": ("客戶編號", "客戶代碼", "代碼", "code"),
    "name": ("客戶名稱", "公司名稱", "名稱", "name"),
    "contact_person": ("聯絡人", "contact_person", "contact"),
    "phone": ("電話", "phone", "tel"),
    "address": ("配送地址", "地址", "address"),
    "tax_id": ("統一編號", "統編", "tax_id"),
    "email": ("電子郵件", "email", "e-mail"),
    "notes": ("備註", "notes"),
}


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def _format_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y/%m/%d")
    return str(value).strip()


def _build_column_map(headers):
    normalized = [_normalize_header(header) for header in headers]
    column_map = {}

    for field_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_key = alias.strip().lower()
            if alias_key in normalized:
                column_map[field_name] = normalized.index(alias_key)
                break

    return column_map


def _phone_columns(headers):
    phone_columns = []
    for index, header in enumerate(headers):
        label = _normalize_header(header)
        if label in {"📞", "電話", "phone", "tel"}:
            phone_columns.append(index)
    return phone_columns


def _cell_value(row, index):
    if index is None or index >= len(row):
        return ""
    return _format_cell(row[index])


def _combine_phones(row, phone_columns):
    phones = []
    for index in phone_columns:
        phone = _cell_value(row, index)
        if phone and phone not in phones:
            phones.append(phone)
    return " / ".join(phones)


def _build_notes(row, headers, column_map):
    note_parts = []
    extra_fields = (
        ("區域", "區域"),
        ("發票地址", "發票地址"),
        ("付款方式", "付款方式"),
        ("固定配送日", "固定配送日"),
        ("配送順序", "配送順序"),
        ("信用額度", "信用額度"),
        ("最後交易日", "最後交易日"),
    )

    for label, prefix in extra_fields:
        normalized = [_normalize_header(header) for header in headers]
        if label.lower() not in normalized:
            continue
        value = _cell_value(row, normalized.index(label.lower()))
        if value:
            note_parts.append(f"{prefix}：{value}")

    notes = _cell_value(row, column_map.get("notes"))
    if notes:
        note_parts.append(notes)

    return "\n".join(note_parts)


def _parse_customer_sheet(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    column_map = _build_column_map(headers)
    phone_columns = _phone_columns(headers)

    if "code" not in column_map or "name" not in column_map:
        raise ValueError("「客戶資料」工作表缺少必要欄位：客戶編號、客戶名稱")

    customers = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        code = _cell_value(row, column_map.get("code"))
        name = _cell_value(row, column_map.get("name"))
        if not code and not name:
            continue

        customers.append(
            {
                "row_number": row_number,
                "code": code,
                "name": name,
                "contact_person": _cell_value(row, column_map.get("contact_person")),
                "phone": _combine_phones(row, phone_columns) or _cell_value(row, column_map.get("phone")),
                "address": _cell_value(row, column_map.get("address")),
                "tax_id": _cell_value(row, column_map.get("tax_id")),
                "email": _cell_value(row, column_map.get("email")),
                "notes": _build_notes(row, headers, column_map),
            }
        )

    return customers


def parse_excel(file_path=None, sheet_name=CUSTOMER_SHEET_NAME):
    path = Path(file_path) if file_path else DEFAULT_EXCEL_PATH
    if not path.exists():
        raise FileNotFoundError(f"找不到 Excel 檔案：{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"找不到工作表：{sheet_name}")

    sheet = workbook[sheet_name]
    customers = _parse_customer_sheet(sheet)
    workbook.close()
    return customers


def import_customers(file_path=None, sheet_name=CUSTOMER_SHEET_NAME):
    rows = parse_excel(file_path, sheet_name=sheet_name)
    result = ImportResult()

    for row in rows:
        row_number = row["row_number"]
        code = row["code"]
        name = row["name"]

        if not code:
            result.skipped += 1
            result.errors.append(f"第 {row_number} 列：缺少客戶編號，已略過")
            continue

        if not name:
            result.skipped += 1
            result.errors.append(f"第 {row_number} 列：缺少客戶名稱，已略過")
            continue

        defaults = {
            "name": name,
            "contact_person": row["contact_person"],
            "phone": row["phone"],
            "address": row["address"],
            "tax_id": row["tax_id"],
            "email": row["email"],
            "notes": row["notes"],
            "is_active": True,
        }

        _, created = Customer.objects.update_or_create(code=code, defaults=defaults)
        if created:
            result.created += 1
        else:
            result.updated += 1

    return result
