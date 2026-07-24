from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook

from apps.inventory.models import Product
from apps.inventory.packaging import parse_packaging_spec

DEFAULT_EXCEL_PATH = settings.BASE_DIR / "data" / "有信ERP.xlsx"
PRODUCT_SHEET = "產品資料"

HEADER_ALIASES = {
    "sku": ("產品編號", "料號"),
    "name": ("產品名稱", "品名"),
    "category": ("產品分類", "產品類型", "分類"),
    "brand": ("品牌",),
    "series": ("系列",),
    "spec": ("規格",),
    "unit": ("單位",),
    "is_for_sale": ("是否販售",),
    "can_be_raw_material": ("可做原料",),
    "is_active": ("啟用",),
    "notes": ("備註",),
}

SALES_UNIT_MAP = {
    "包": Product.SalesUnit.PACK,
    "罐": Product.SalesUnit.CAN,
    "袋": Product.SalesUnit.BAG,
    "箱": Product.SalesUnit.BOX,
    "box": Product.SalesUnit.BOX,
    "kg": Product.SalesUnit.KG,
    "公斤": Product.SalesUnit.KG,
}

INVENTORY_UNIT_MAP = {
    "包": Product.Unit.PCS,
    "罐": Product.Unit.PCS,
    "袋": Product.Unit.PCS,
    "箱": Product.Unit.BOX,
    "box": Product.Unit.BOX,
    "kg": Product.Unit.KG,
    "公斤": Product.Unit.KG,
    "g": Product.Unit.G,
    "公克": Product.Unit.G,
    "l": Product.Unit.L,
    "公升": Product.Unit.L,
    "ml": Product.Unit.ML,
    "毫升": Product.Unit.ML,
    "pcs": Product.Unit.PCS,
    "件": Product.Unit.PCS,
}


@dataclass
class ProductImportReport:
    total_rows: int = 0
    success_count: int = 0
    failed_count: int = 0
    duplicate_skus: list[str] = field(default_factory=list)
    failures: list[dict] = field(default_factory=list)
    successes: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return not self.errors


def _norm(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (datetime, date)):
        return value
    text = str(value).strip()
    return text if text else None


def _checkmark(value) -> bool:
    return _norm(value) in {"✓", "v", "V", "Y", "yes", "是", "1", "True", "true"}


def _uncheckmark(value) -> bool:
    normalized = _norm(value)
    if normalized is None:
        return False
    return normalized in {"✗", "x", "X", "N", "no", "否", "0", "False", "false"}


def _bool_from_sheet(value, *, default=True) -> bool:
    normalized = _norm(value)
    if normalized is None:
        return default
    if _checkmark(value):
        return True
    if _uncheckmark(value):
        return False
    return default


def _header_map(header_row) -> dict[str, int]:
    mapping = {}
    for idx, cell in enumerate(header_row):
        label = _norm(cell)
        if not label:
            continue
        for key, aliases in HEADER_ALIASES.items():
            if label in aliases and key not in mapping:
                mapping[key] = idx
    return mapping


def _cell(row, index: int | None):
    if index is None:
        return None
    if index >= len(row):
        return None
    return row[index]


def _map_sales_unit(value) -> str:
    label = (_norm(value) or "").lower()
    if not label:
        return Product.SalesUnit.PACK
    return SALES_UNIT_MAP.get(label, SALES_UNIT_MAP.get(_norm(value), Product.SalesUnit.PACK))


def _map_inventory_unit(value) -> str:
    label = (_norm(value) or "").lower()
    if not label:
        return Product.Unit.PCS
    return INVENTORY_UNIT_MAP.get(label, INVENTORY_UNIT_MAP.get(_norm(value), Product.Unit.PCS))


def _packaging_from_row(*, spec: str, unit_label: str) -> dict:
    parsed = parse_packaging_spec(spec)
    data = {"sales_unit": _map_sales_unit(unit_label)}
    if parsed:
        data["net_weight_value"] = parsed.get("net_weight_value")
        data["net_weight_unit"] = parsed.get("net_weight_unit", "")
        if parsed.get("sales_unit"):
            data["sales_unit"] = parsed["sales_unit"]
    return data


def parse_product_sheet(sheet) -> ProductImportReport:
    report = ProductImportReport()
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        report.errors.append(f"「{PRODUCT_SHEET}」工作表為空")
        return report

    headers = _header_map(rows[0])
    required = ("sku", "name")
    missing = [key for key in required if key not in headers]
    if missing:
        report.errors.append(f"缺少必要欄位：{', '.join(missing)}")
        return report

    parsed_rows = []
    seen_skus: dict[str, int] = {}

    for row_number, row in enumerate(rows[1:], start=2):
        cells = list(row)
        sku = _norm(_cell(cells, headers.get("sku")))
        name = _norm(_cell(cells, headers.get("name")))
        if not sku and not name:
            continue

        report.total_rows += 1

        if not sku:
            report.failed_count += 1
            report.failures.append({"row": row_number, "sku": "", "reason": "缺少產品編號"})
            continue
        if not name:
            report.failed_count += 1
            report.failures.append({"row": row_number, "sku": sku, "reason": "缺少產品名稱"})
            continue

        if sku in seen_skus:
            if sku not in report.duplicate_skus:
                report.duplicate_skus.append(sku)
            report.failed_count += 1
            report.failures.append(
                {
                    "row": row_number,
                    "sku": sku,
                    "reason": f"產品編號重複（首見於第 {seen_skus[sku]} 列）",
                }
            )
            continue
        seen_skus[sku] = row_number

        spec = _norm(_cell(cells, headers.get("spec"))) or ""
        unit_label = _norm(_cell(cells, headers.get("unit"))) or "包"
        can_be_raw = _bool_from_sheet(_cell(cells, headers.get("can_be_raw_material")), default=False)

        parsed_rows.append(
            {
                "row_number": row_number,
                "sku": sku,
                "name": name,
                "category": _norm(_cell(cells, headers.get("category"))) or "",
                "brand": _norm(_cell(cells, headers.get("brand"))) or "",
                "series": _norm(_cell(cells, headers.get("series"))) or "",
                "spec": spec,
                "unit_label": unit_label,
                "is_for_sale": _bool_from_sheet(_cell(cells, headers.get("is_for_sale")), default=True),
                "can_be_raw_material": can_be_raw,
                "is_active": _bool_from_sheet(_cell(cells, headers.get("is_active")), default=True),
                "notes": _norm(_cell(cells, headers.get("notes"))) or "",
                "product_kind": Product.ProductKind.DUAL if can_be_raw else Product.ProductKind.FINISHED,
                "packaging": _packaging_from_row(spec=spec, unit_label=unit_label),
            }
        )

    report._parsed_rows = parsed_rows
    return report


def run_preflight(excel_path: Path | None = None) -> ProductImportReport:
    path = Path(excel_path) if excel_path else DEFAULT_EXCEL_PATH
    report = ProductImportReport()

    if not path.exists():
        report.errors.append(f"找不到 Excel 檔案：{path}")
        return report

    wb = load_workbook(path, read_only=True, data_only=True)
    if PRODUCT_SHEET not in wb.sheetnames:
        report.errors.append(f"找不到工作表：{PRODUCT_SHEET}")
        return report

    return parse_product_sheet(wb[PRODUCT_SHEET])


def execute_import(report: ProductImportReport) -> ProductImportReport:
    if not report.passed:
        return report

    parsed_rows = getattr(report, "_parsed_rows", None)
    if parsed_rows is None:
        report.errors.append("請先執行 preflight")
        return report

    for row in parsed_rows:
        try:
            existing = Product.objects.filter(sku=row["sku"]).first()
            defaults = {
                "name": row["name"],
                "category": row["category"],
                "brand": row["brand"],
                "series": row["series"],
                "spec": row["spec"],
                "product_kind": row["product_kind"],
                "is_for_sale": row["is_for_sale"],
                "can_be_raw_material": row["can_be_raw_material"],
                "is_active": row["is_active"],
                "description": row["notes"],
                "sales_unit": row["packaging"]["sales_unit"],
                "net_weight_value": row["packaging"].get("net_weight_value"),
                "net_weight_unit": row["packaging"].get("net_weight_unit", ""),
            }
            if existing is None:
                defaults["unit"] = _map_inventory_unit(row["unit_label"])

            product, created = Product.objects.update_or_create(
                sku=row["sku"],
                defaults=defaults,
            )
            report.success_count += 1
            report.successes.append(
                {
                    "row": row["row_number"],
                    "sku": row["sku"],
                    "name": row["name"],
                    "action": "created" if created else "updated",
                    "packaging": product.packaging_display,
                }
            )
        except Exception as exc:
            report.failed_count += 1
            report.failures.append(
                {"row": row["row_number"], "sku": row["sku"], "reason": str(exc)}
            )

    return report
