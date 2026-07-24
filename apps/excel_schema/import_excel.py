from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import django
from django.apps import apps
from django.db import transaction
from openpyxl import load_workbook

DEFAULT_EXCEL_PATH = Path("/Users/zhengcongsheng/Downloads/有信ERP.xlsx")

SHEET_ORDER = [
    "首頁",
    "今日配送",
    "快速查詢",
    "接單中心",
    "收款管理",
    "客戶資料",
    "產品資料",
    "原料資料",
    "配方管理",
    "配方主檔",
    "配方明細",
    "生產記錄",
    "採購進貨",
    "庫存管理",
    "供應商資料",
    "配送規則",
    "價格管理",
    "廠商資料",
    "AI控制台",
    "系統設定",
    "改善中心",
]


@dataclass
class SheetImportResult:
    sheet_name: str
    imported: int = 0
    skipped: int = 0
    success: bool = True
    errors: list[str] = field(default_factory=list)


def setup_django():
    project_root = Path(__file__).resolve().parents[2]
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
    django.setup()


def _cell_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _is_blank(value):
    return value is None or str(value).strip() == ""


def _is_empty_row(row):
    if not row:
        return True
    return all(_is_blank(cell) for cell in row)


def _parse_decimal(value):
    if _is_blank(value):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _parse_int(value):
    if _is_blank(value):
        return None
    try:
        if isinstance(value, float):
            return int(value)
        return int(str(value).strip())
    except (ValueError, TypeError):
        return None


def _parse_date(value):
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value):
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return None


def _get_models():
    from apps.excel_schema import models as m

    return m


def _clear_all_tables():
    models = _get_models()
    delete_order = [
        models.SheetRecipeDetail,
        models.SheetRecipeMaster,
        models.SheetRecipeManagement,
        models.SheetPurchaseReceipt,
        models.SheetDeliveryRule,
        models.SheetTodayDelivery,
        models.SheetQuickSearch,
        models.SheetOrderCenter,
        models.SheetCustomerData,
        models.SheetProductData,
        models.SheetRawMaterialData,
        models.SheetHomepage,
        models.SheetPaymentManagement,
        models.SheetProductionRecord,
        models.SheetInventoryManagement,
        models.SheetSupplierData,
        models.SheetPriceManagement,
        models.SheetVendorData,
        models.SheetAiConsole,
        models.SheetSystemSettings,
        models.SheetImprovementCenter,
    ]
    for model in delete_order:
        model.objects.all().delete()


def _lookup_customer_by_code(code):
    if _is_blank(code):
        return None
    from apps.excel_schema.models import SheetCustomerData

    return SheetCustomerData.objects.filter(customer_code=_cell_str(code)).first()


def _lookup_customer_by_name(name):
    if _is_blank(name):
        return None
    from apps.excel_schema.models import SheetCustomerData

    return SheetCustomerData.objects.filter(customer_name=_cell_str(name)).first()


def _lookup_product_by_code(code):
    if _is_blank(code):
        return None
    from apps.excel_schema.models import SheetProductData

    return SheetProductData.objects.filter(product_code=_cell_str(code)).first()


def _lookup_product_by_name(name):
    if _is_blank(name):
        return None
    from apps.excel_schema.models import SheetProductData

    return SheetProductData.objects.filter(product_name=_cell_str(name)).first()


def _lookup_material_by_code(code):
    if _is_blank(code):
        return None
    from apps.excel_schema.models import SheetRawMaterialData

    return SheetRawMaterialData.objects.filter(material_code=_cell_str(code)).first()


def _lookup_recipe_by_id(recipe_id):
    if _is_blank(recipe_id):
        return None
    from apps.excel_schema.models import SheetRecipeMaster

    return SheetRecipeMaster.objects.filter(recipe_id=_cell_str(recipe_id)).first()


def _has_any_data(row, limit=20):
    if not row:
        return False
    for cell in row[:limit]:
        if not _is_blank(cell):
            return True
    return False


def _import_rows_with_savepoints(rows, import_row, result):
    for row_number, row in rows:
        if not _has_any_data(row):
            continue
        try:
            with transaction.atomic():
                imported = import_row(row, row_number)
                if imported:
                    result.imported += 1
                else:
                    result.skipped += 1
        except Exception as exc:
            result.skipped += 1
            result.errors.append(f"第 {row_number} 列：{exc}")


def _row_value(row, index):
    if index >= len(row):
        return None
    return row[index]


def import_sheet_homepage(ws):
    from apps.excel_schema.models import SheetHomepage

    rows = list(ws.iter_rows(values_only=True))
    row_map = {}
    for row in rows:
        if not row or _is_blank(row[0]):
            continue
        label = _cell_str(row[0])
        value = row[3] if len(row) > 3 else None
        if _is_blank(value) and len(row) > 1:
            value = row[1]
        row_map[label] = value

    record = SheetHomepage(
        stat_today_delivery=_cell_str(row_map.get("🚚 今日配送", "")),
        stat_today_payment=_cell_str(row_map.get("💰 今日收款", "")),
        stat_today_purchase=_cell_str(row_map.get("📦 今日採購", "")),
        stat_today_reminder=_cell_str(row_map.get("⚠ 今日提醒", "")),
        ai_suggestion_1=_cell_str(rows[19][0]) if len(rows) > 19 and rows[19] else "",
        ai_suggestion_2=_cell_str(rows[20][0]) if len(rows) > 20 and rows[20] else "",
        ai_suggestion_3=_cell_str(rows[21][0]) if len(rows) > 21 and rows[21] else "",
    )
    record.save()
    return SheetImportResult(sheet_name="首頁", imported=1)


def import_sheet_today_delivery(ws):
    from apps.excel_schema.models import SheetTodayDelivery

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return SheetImportResult(sheet_name="今日配送")

    result = SheetImportResult(sheet_name="今日配送")

    def import_row(row, row_number):
        customer_name = _cell_str(_row_value(row, 1))
        product_name = _cell_str(_row_value(row, 4))
        customer = _lookup_customer_by_name(customer_name) if customer_name else None
        product = _lookup_product_by_name(product_name) if product_name else None

        SheetTodayDelivery.objects.create(
            delivery_date=_parse_date(_row_value(row, 0)),
            customer=customer,
            delivery_area=_cell_str(_row_value(row, 2)),
            address=_cell_str(_row_value(row, 3)),
            product=product,
            quantity=_parse_decimal(_row_value(row, 5)),
            delivery_sequence=_parse_int(_row_value(row, 6)),
            status=_cell_str(_row_value(row, 7)),
            completed_at=_parse_datetime(_row_value(row,  8)),
            notes=_cell_str(_row_value(row, 9)),
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2)],
        import_row,
        result,
    )
    return result


def import_sheet_quick_search(ws):
    from apps.excel_schema.models import SheetQuickSearch

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="快速查詢")
    search_keyword = _cell_str(_row_value(rows[2], 1)) if len(rows) > 2 else ""

    header_row_index = None
    for index, row in enumerate(rows):
        if row and _cell_str(_row_value(row, 0)) == "客戶編號":
            header_row_index = index
            break

    if header_row_index is None:
        return result

    data_rows = [
        (row_number, row)
        for row_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2)
        if _has_any_data(row)
    ]

    def import_row(row, row_number):
        customer_code = _cell_str(_row_value(row, 0))
        if not customer_code:
            return False
        customer = _lookup_customer_by_code(customer_code)
        SheetQuickSearch.objects.create(
            search_keyword=search_keyword,
            customer=customer,
            region=_cell_str(_row_value(row, 1)),
            customer_name=_cell_str(_row_value(row, 2)),
            contact_person=_cell_str(_row_value(row, 3)),
            phone_1=_cell_str(_row_value(row, 4)),
            phone_2=_cell_str(_row_value(row, 5)),
            phone_3=_cell_str(_row_value(row, 6)),
            map_address=_cell_str(_row_value(row, 7)),
            create_order_action=_cell_str(_row_value(row, 8)),
        )
        return True

    _import_rows_with_savepoints(data_rows, import_row, result)
    return result


def import_sheet_order_center(ws):
    from apps.excel_schema.models import SheetOrderCenter

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="接單中心")

    search_input = _cell_str(_row_value(rows[3], 0)) if len(rows) > 3 else ""
    kv = {}
    for row in rows:
        if row and not _is_blank(row[0]) and len(row) > 1 and not _is_blank(row[1]):
            if _cell_str(row[0]) not in {"建立訂單", "📋 歷史訂單", "品名"}:
                kv[_cell_str(row[0])] = row[1]

    profile = SheetOrderCenter(
        search_input=search_input,
        customer_name=_cell_str(kv.get("客戶姓名", "")),
        phone_1=_cell_str(kv.get("電話", "")),
        address_1=_cell_str(kv.get("地址", "")),
        payment_method_1=_cell_str(kv.get("付款方式", "")),
    )
    profile.save()
    result.imported += 1

    order_date = _parse_date(kv.get("訂單日期"))
    order_customer = _cell_str(kv.get("客戶", ""))
    order_phone = ""
    order_address = ""
    order_payment = ""
    order_kv_started = False
    for row in rows:
        if row and _cell_str(_row_value(row, 0)) == "訂單日期":
            order_kv_started = True
            continue
        if order_kv_started and row and not _is_blank(row[0]) and len(row) > 1:
            label = _cell_str(row[0])
            if label == "電話":
                order_phone = _cell_str(row[1])
            elif label == "地址":
                order_address = _cell_str(row[1])
            elif label == "付款方式":
                order_payment = _cell_str(row[1])

    for row in rows:
        if not row or _is_blank(_row_value(row, 0)):
            continue
        first = _cell_str(_row_value(row, 0))
        if first in {"中西區", "北區", "南區", "東區", "安平區", "永康區", "新化區", "玉井區", "茄萣區", "歸仁區"} or (
            len(row) > 3 and not _is_blank(_row_value(row, 1)) and _is_blank(_row_value(row, 2)) and not _is_blank(_row_value(row, 3))
        ):
            if first in {"客戶姓名", "電話", "地址", "付款方式", "訂單日期", "客戶", "品名", "數量", "單位"}:
                continue
            SheetOrderCenter.objects.create(
                search_input=search_input,
                region=first if first.endswith("區") else "",
                customer_name=_cell_str(_row_value(row, 1)),
                phone_1=_cell_str(_row_value(row, 3)),
            )
            result.imported += 1

    line_header_index = None
    for index, row in enumerate(rows):
        if row and _cell_str(_row_value(row, 0)) == "品名":
            line_header_index = index
            break

    if line_header_index is not None:
        for row in rows[line_header_index + 1 :]:
            if not _has_any_data(row):
                continue
            quantity = _parse_decimal(_row_value(row, 1))
            unit = _cell_str(_row_value(row, 2))
            product_name = _cell_str(_row_value(row, 0))
            if not product_name and quantity is None and not unit:
                continue
            SheetOrderCenter.objects.create(
                search_input=search_input,
                order_date=order_date,
                order_customer=order_customer,
                phone_2=order_phone,
                address_2=order_address,
                payment_method_2=order_payment,
                product_name=product_name,
                quantity=quantity,
                unit=unit,
            )
            result.imported += 1

    return result


def import_sheet_empty(sheet_name):
    return SheetImportResult(sheet_name=sheet_name)


def import_sheet_customer_data(ws):
    from apps.excel_schema.models import SheetCustomerData

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="客戶資料")

    def import_row(row, row_number):
        customer_code = _cell_str(_row_value(row, 0))
        customer_name = _cell_str(_row_value(row, 2))
        if not customer_code or not customer_name:
            return False

        SheetCustomerData.objects.update_or_create(
            customer_code=customer_code,
            defaults={
                "region": _cell_str(_row_value(row, 1)),
                "customer_name": customer_name,
                "contact_person": _cell_str(_row_value(row, 3)),
                "phone_1": _cell_str(_row_value(row, 4)),
                "phone_2": _cell_str(_row_value(row, 5)),
                "phone_3": _cell_str(_row_value(row, 6)),
                "delivery_address": _cell_str(_row_value(row, 7)),
                "invoice_address": _cell_str(_row_value(row, 8)),
                "map_link": _cell_str(_row_value(row, 9)),
                "line_id": _cell_str(_row_value(row, 10)),
                "payment_method": _cell_str(_row_value(row, 11)),
                "fixed_delivery_day": _cell_str(_row_value(row, 12)),
                "delivery_sequence": _parse_int(_row_value(row, 13)),
                "credit_limit": _parse_decimal(_row_value(row, 14)),
                "last_transaction_date": _parse_date(_row_value(row, 15)),
                "notes": _cell_str(_row_value(row, 16)),
            },
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


def import_sheet_product_data(ws):
    from apps.excel_schema.models import SheetProductData

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="產品資料")

    def import_row(row, row_number):
        product_code = _cell_str(_row_value(row, 0))
        product_name = _cell_str(_row_value(row, 1))
        if not product_code or not product_name:
            return False

        SheetProductData.objects.update_or_create(
            product_code=product_code,
            defaults={
                "product_name": product_name,
                "product_type": _cell_str(_row_value(row, 2)),
                "specification": _cell_str(_row_value(row, 3)),
                "unit": _cell_str(_row_value(row, 4)),
                "is_for_sale": _cell_str(_row_value(row, 5)),
                "can_be_raw_material": _cell_str(_row_value(row, 6)),
                "is_active": _cell_str(_row_value(row, 7)),
                "notes": _cell_str(_row_value(row, 8)),
                "cost": _parse_decimal(_row_value(row, 9)),
                "price": _parse_decimal(_row_value(row, 10)),
            },
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


def import_sheet_raw_material_data(ws):
    from apps.excel_schema.models import SheetRawMaterialData

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="原料資料")

    def import_row(row, row_number):
        material_code = _cell_str(_row_value(row, 0))
        material_name = _cell_str(_row_value(row, 1))
        if not material_code or not material_name:
            return False

        SheetRawMaterialData.objects.update_or_create(
            material_code=material_code,
            defaults={
                "material_name": material_name,
                "category": _cell_str(_row_value(row, 2)),
                "unit": _cell_str(_row_value(row, 3)),
                "cost_per_kg": _parse_decimal(_row_value(row, 4)),
                "latest_purchase_price": _parse_decimal(_row_value(row, 5)),
                "last_purchase_date": _parse_date(_row_value(row, 6)),
                "safety_stock": _parse_decimal(_row_value(row, 7)),
                "is_active": _cell_str(_row_value(row, 8)),
                "notes": _cell_str(_row_value(row, 9)),
            },
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


def import_sheet_recipe_management(ws):
    from apps.excel_schema.models import SheetRecipeManagement

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="配方管理")

    def import_row(row, row_number):
        product_code = _cell_str(_row_value(row, 0))
        material_code = _cell_str(_row_value(row, 3))
        if not product_code or not material_code:
            return False

        product = _lookup_product_by_code(product_code)
        material = _lookup_material_by_code(material_code)
        if not product or not material:
            raise ValueError("找不到產品或原料 FK")

        SheetRecipeManagement.objects.create(
            product=product,
            product_name=_cell_str(_row_value(row, 1)),
            version=_cell_str(_row_value(row, 2)),
            material=material,
            material_name=_cell_str(_row_value(row, 4)),
            quantity=_parse_decimal(_row_value(row, 5)),
            unit=_cell_str(_row_value(row, 6)),
            sort_order=_parse_int(_row_value(row, 7)),
            is_active=_cell_str(_row_value(row, 8)),
            notes=_cell_str(_row_value(row, 9)),
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


def import_sheet_recipe_master(ws):
    from apps.excel_schema.models import SheetRecipeMaster

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="配方主檔")

    def import_row(row, row_number):
        recipe_id = _cell_str(_row_value(row, 0))
        product_code = _cell_str(_row_value(row, 1))
        if not recipe_id or not product_code:
            return False

        product = _lookup_product_by_code(product_code)
        if not product:
            raise ValueError("找不到產品編號 FK")

        SheetRecipeMaster.objects.update_or_create(
            recipe_id=recipe_id,
            defaults={
                "product": product,
                "product_name": _cell_str(_row_value(row, 2)),
                "product_type": _cell_str(_row_value(row, 3)),
                "product_feature": _cell_str(_row_value(row, 4)),
                "recipe_version": _cell_str(_row_value(row, 5)),
                "base_batch_kg": _parse_decimal(_row_value(row, 6)),
                "status": _cell_str(_row_value(row, 7)),
                "created_date": _parse_date(_row_value(row, 8)),
                "change_reason": _cell_str(_row_value(row, 9)),
                "notes": _cell_str(_row_value(row, 10)),
                "total_cost": _parse_decimal(_row_value(row, 11)),
                "cost_per_kg": _parse_decimal(_row_value(row, 12)),
            },
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


def import_sheet_recipe_detail(ws):
    from apps.excel_schema.models import SheetRecipeDetail

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="配方明細")

    def import_row(row, row_number):
        recipe_id = _cell_str(_row_value(row, 0))
        item_no = _parse_int(_row_value(row, 1))
        material_code = _cell_str(_row_value(row, 2))
        if not recipe_id or item_no is None or not material_code:
            return False

        recipe = _lookup_recipe_by_id(recipe_id)
        material = _lookup_material_by_code(material_code)
        if not recipe or not material:
            raise ValueError("找不到配方或原料 FK")

        SheetRecipeDetail.objects.update_or_create(
            recipe=recipe,
            item_no=item_no,
            defaults={
                "material": material,
                "material_name": _cell_str(_row_value(row, 3)),
                "selected_item": _cell_str(_row_value(row, 4)),
                "quantity": _parse_decimal(_row_value(row, 5)),
                "unit": _cell_str(_row_value(row, 6)),
                "notes": _cell_str(_row_value(row, 7)),
                "cost_per_kg": _parse_decimal(_row_value(row, 8)),
                "cost": _parse_decimal(_row_value(row, 9)),
            },
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


def import_sheet_purchase_receipt(ws):
    from apps.excel_schema.models import SheetPurchaseReceipt

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="採購進貨")

    def import_row(row, row_number):
        receipt_no = _cell_str(_row_value(row, 0))
        material_code = _cell_str(_row_value(row, 2))
        if not receipt_no or not material_code:
            return False

        material = _lookup_material_by_code(material_code)
        if not material:
            raise ValueError("找不到原料編號 FK")

        SheetPurchaseReceipt.objects.update_or_create(
            receipt_no=receipt_no,
            material=material,
            defaults={
                "receipt_date": _parse_date(_row_value(row, 1)),
                "material_name": _cell_str(_row_value(row, 3)),
                "quantity": _parse_decimal(_row_value(row, 4)),
                "unit_price": _parse_decimal(_row_value(row, 5)),
                "supplier_name": _cell_str(_row_value(row, 6)),
                "notes": _cell_str(_row_value(row, 7)),
            },
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


def import_sheet_delivery_rule(ws):
    from apps.excel_schema.models import SheetDeliveryRule

    rows = list(ws.iter_rows(values_only=True))
    result = SheetImportResult(sheet_name="配送規則")

    def import_row(row, row_number):
        customer_code = _cell_str(_row_value(row, 0))
        if not customer_code:
            return False

        customer = _lookup_customer_by_code(customer_code)
        if not customer:
            raise ValueError("找不到客戶編號 FK")

        fixed_product = _lookup_product_by_code(_row_value(row, 5))
        SheetDeliveryRule.objects.create(
            customer=customer,
            customer_name=_cell_str(_row_value(row, 1)),
            weekday=_cell_str(_row_value(row, 2)),
            delivery_area=_cell_str(_row_value(row, 3)),
            delivery_sequence=_parse_int(_row_value(row, 4)),
            fixed_product=fixed_product,
            fixed_quantity=_parse_decimal(_row_value(row, 6)),
            is_delivery=_cell_str(_row_value(row, 7)),
            notes=_cell_str(_row_value(row, 8)),
            created_date=_parse_date(_row_value(row, 9)),
            updated_at=_parse_datetime(_row_value(row, 10)),
        )
        return True

    _import_rows_with_savepoints(
        [(row_number, row) for row_number, row in enumerate(rows[1:], start=2) if _has_any_data(row)],
        import_row,
        result,
    )
    return result


SHEET_IMPORTERS = {
    "首頁": import_sheet_homepage,
    "今日配送": import_sheet_today_delivery,
    "快速查詢": import_sheet_quick_search,
    "接單中心": import_sheet_order_center,
    "收款管理": lambda ws: import_sheet_empty("收款管理"),
    "客戶資料": import_sheet_customer_data,
    "產品資料": import_sheet_product_data,
    "原料資料": import_sheet_raw_material_data,
    "配方管理": import_sheet_recipe_management,
    "配方主檔": import_sheet_recipe_master,
    "配方明細": import_sheet_recipe_detail,
    "生產記錄": lambda ws: import_sheet_empty("生產記錄"),
    "採購進貨": import_sheet_purchase_receipt,
    "庫存管理": lambda ws: import_sheet_empty("庫存管理"),
    "供應商資料": lambda ws: import_sheet_empty("供應商資料"),
    "配送規則": import_sheet_delivery_rule,
    "價格管理": lambda ws: import_sheet_empty("價格管理"),
    "廠商資料": lambda ws: import_sheet_empty("廠商資料"),
    "AI控制台": lambda ws: import_sheet_empty("AI控制台"),
    "系統設定": lambda ws: import_sheet_empty("系統設定"),
    "改善中心": lambda ws: import_sheet_empty("改善中心"),
}


def import_sheet(workbook, sheet_name):
    ws = workbook[sheet_name]
    importer = SHEET_IMPORTERS[sheet_name]
    try:
        with transaction.atomic():
            return importer(ws)
    except Exception as exc:
        return SheetImportResult(
            sheet_name=sheet_name,
            success=False,
            errors=[str(exc)],
        )


def print_table_counts():
    models = _get_models()
    table_models = [
        models.SheetHomepage,
        models.SheetTodayDelivery,
        models.SheetQuickSearch,
        models.SheetOrderCenter,
        models.SheetPaymentManagement,
        models.SheetCustomerData,
        models.SheetProductData,
        models.SheetRawMaterialData,
        models.SheetRecipeManagement,
        models.SheetRecipeMaster,
        models.SheetRecipeDetail,
        models.SheetProductionRecord,
        models.SheetPurchaseReceipt,
        models.SheetInventoryManagement,
        models.SheetSupplierData,
        models.SheetDeliveryRule,
        models.SheetPriceManagement,
        models.SheetVendorData,
        models.SheetAiConsole,
        models.SheetSystemSettings,
        models.SheetImprovementCenter,
    ]
    print("\n=== 資料表匯入結果 ===")
    for model in table_models:
        print(f"{model._meta.db_table}: {model.objects.count()} 筆")


def run_import(excel_path=DEFAULT_EXCEL_PATH):
    setup_django()
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 Excel 檔案：{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    _clear_all_tables()

    results = []
    for sheet_name in SHEET_ORDER:
        if sheet_name not in workbook.sheetnames:
            result = SheetImportResult(
                sheet_name=sheet_name,
                success=False,
                errors=[f"工作表不存在：{sheet_name}"],
            )
        else:
            result = import_sheet(workbook, sheet_name)

        results.append(result)
        status = "成功" if result.success else "失敗"
        print(f"[{sheet_name}] 匯入 {result.imported} 筆 | 跳過 {result.skipped} 筆 | {status}")
        for error in result.errors:
            print(f"  - {error}")

    workbook.close()
    print_table_counts()
    return results


if __name__ == "__main__":
    run_import()
